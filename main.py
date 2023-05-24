import tkinter as tk
from tkinter import ttk
import openpyxl
import paramiko
import threading
import json
import os

USERNAME = "user" # credidentiale temporare
PASSWORD = "parola"

workbook = openpyxl.load_workbook('vms.xlsx')
worksheet = workbook['Sheet1']

window = tk.Tk()
window.title("Agent install")
window.geometry("700x400")

# Search bar and search button
search_frame = tk.Frame(window)
search_frame.pack(side=tk.TOP, fill=tk.X)

search_label = tk.Label(search_frame, text="Cauta VM:")
search_label.grid(row=0, column=0, padx=5, pady=5, sticky='n')

search_bar = tk.Entry(search_frame, width=50)
search_bar.grid(row=0, column=1, padx=5, pady=5, sticky='n')

def search():
    nume = search_bar.get()
    found = False
    clear_frame(data_frame)
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == nume:
            found = True
            topologie = worksheet.cell(row=row, column=2).value
            owner = worksheet.cell(row=row, column=3).value
            vm = worksheet.cell(row=row, column=4).value
            Mplane = worksheet.cell(row=row, column=5).value

            nume_label = tk.Label(data_frame, text=f"Name: {nume}")
            nume_label.pack()

            topologie_label = tk.Label(data_frame, text=f"Topologie: {topologie}")
            topologie_label.pack()

            owner_label = tk.Label(data_frame, text=f"Owner: {owner}")
            owner_label.pack()

            vm_label = tk.Label(data_frame, text=f"VM: {vm}")
            vm_label.pack()

            Mplane_label = tk.Label(data_frame, text=f"Mplane: {Mplane}")
            Mplane_label.pack()

            # SSH button for the VM
            ssh_button = tk.Button(data_frame, text="Connect SSH", command=lambda vm=vm: connect_ssh_thread(vm))
            ssh_button.pack(pady=5)
            agent_buttons(vm)
            disable_buttons(data_frame, vm)
    ssh_status_label.config(text="")
    search_not_found(found,nume)

def search_not_found(found,nume):
    if not found:
        clear_frame(data_frame)
        text_label = tk.Label(data_frame, text=f"Nu s-a gasit {nume}.",fg="red")
        text_label.pack()

def clear_frame(frame):
    for widget in frame.winfo_children():
        widget.destroy()

def agent_buttons(vm):
    ute_ca_button(vm)
    agent_gve_common_button(vm)

def ute_ca_button(vm):
    with open('agent_ute-ca.json', 'r') as f:
        data = json.load(f)
    ute_ca_button = tk.Button(data_frame, text=f"UTE_CA - {vm}", command=lambda: [os.system(f'sshpass -p {PASSWORD} ssh {USERNAME}@{vm} "{cmd}"') for cmd in data])
    ute_ca_button.pack(pady=5)

def agent_gve_common_button(vm):
    with open('agent_gve_common.json', 'r') as f:
        data = json.load(f)
    agent_gve_common_button = tk.Button(data_frame, text=f"AGENT_GVE_COMMON - {vm}", command=lambda: [os.system(f'sshpass -p {PASSWORD} ssh {USERNAME}@{vm} "{cmd}"') for cmd in data])
    agent_gve_common_button.pack(pady=5)

def enable_buttons(frame, vm):
    for widget in frame.winfo_children():
        if isinstance(widget, tk.Button) and widget["text"] == f"UTE_CA - {vm}" or widget["text"] == f"AGENT_GVE_COMMON - {vm}":
            widget["state"] = "normal"

def disable_buttons(frame, vm):
    for widget in frame.winfo_children():
        if isinstance(widget, tk.Button) and widget["text"] == f"UTE_CA - {vm}" or widget["text"] == f"AGENT_GVE_COMMON - {vm}":
            widget["state"] = "disabled"

def connect_ssh_thread(vm):
    thread = threading.Thread(target=ssh_connect, args=(vm,))
    thread.start()

def ssh_connect(vm):
    # SSH client object
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(hostname=vm, username=USERNAME, password=PASSWORD)
        ssh_status_label.config(text=f"Conexiune SSH reusita pentru {vm}", fg="green")
        enable_buttons(data_frame, vm)
        quit_button = tk.Button(window, text="Quit SSH", command=lambda: quit_ssh(ssh, vm))
        quit_button.pack(pady=5,side=tk.RIGHT, anchor=tk.CENTER)
        first_ssh_button = tk.Button(window, text="First SSH", command=lambda: os.system(f'ssh-keyscan -H {vm} >> ~/.ssh/known_hosts && ssh {USERNAME}@{vm} echo "Conexiune SSH reusita"'))
        first_ssh_button.pack(pady=5,side=tk.RIGHT, anchor=tk.CENTER)
        def quit_ssh(ssh, vm):
            ssh.close()
            ssh_status_label.config(text=f"Conexiune SSH inchisa pentru {vm}", fg="red")
            disable_buttons(data_frame, vm)
            first_ssh_button.destroy()
            quit_button.destroy()
    except paramiko.AuthenticationException:
        # Displays an authentication error:
        print("Eroare de autentificare. Verifica username-ul si parola.")
        ssh_status_label.config(text=f"Conexiune esuata pentru {vm}. Verifica username-ul si parola.", fg="red")
        disable_buttons(data_frame, vm)
    except paramiko.SSHException:
        # Displays any other SSH-related error:
        print("Eroare SSH. Verifica conexiunea la retea.")
        ssh_status_label.config(text=f"Conexiune esuata pentru {vm}. Verifica conexiunea la retea.", fg="red")
        disable_buttons(data_frame, vm)
    except Exception as e:
        # Displays any other error:
        print("Unexpected error:", e)
        ssh_status_label.config(text=f"Conexiune esuata pentru {vm}. Unexpected error {e}.", fg="red")
        disable_buttons(data_frame, vm)

search_button = tk.Button(search_frame, text="Cauta", command=search)
search_button.grid(row=0, column=2, padx=5, pady=5, sticky='n')

first_ssh_label = tk.Label(search_frame, text="Folositi butonul First SSH doar la prima conectare la un VM!", fg="red")
first_ssh_label.grid(row=1, columnspan=3, padx=5, pady=5, sticky='n')

ssh_status_label = tk.Label(search_frame, text="")
ssh_status_label.grid(row=2, columnspan=3, padx=5, pady=5, sticky='n')

search_frame.columnconfigure(1, weight=1)

# Canvas widget with scrollbar
canvas = tk.Canvas(window)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

scrollbar = ttk.Scrollbar(window, orient=tk.VERTICAL, command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))

# Frame to hold the data
data_frame = tk.Frame(canvas)

# Add the frame to the Canvas widget
canvas.create_window((0, 0), window=data_frame, anchor='nw')

# Run the Tkinter event loop
window.mainloop()